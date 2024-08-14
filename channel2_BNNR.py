import numpy as np
import openpyxl
import pandas as pd
# write execl data into an array
def read(exc_path):
    wb = openpyxl.load_workbook(exc_path)
    data = wb.active
    data_row = data.max_row
    data_col = data.max_column
    num = np.zeros((data_row, data_col))
    for i in range(data_row):
        for j in range(data_col):
            num[i][j] = data.cell(i + 1, j + 1).value
    return num
# SVT: singular value thresholding operator for matrix Y by thretholding parameter x
def svt(Y, x):
    S, V, D = np.linalg.svd(Y)
    for index in range(len(V)):
        if V[index] >= x:
            V[index] = V[index] - x
        else:
            V[index] = 0
    v = np.diag(V)
    V_row = np.array(Y).shape[0]
    V_col = np.array(Y).shape[1]
    if V_row < V_col:
        v_new = np.hstack((v, np.zeros((V_row, V_col - V_row))))
    else:
        v_new = np.vstack((v, np.zeros((V_row-V_col, V_col))))
    e = np.dot(S, np.dot(v_new, D))
    return e
def BNNR(alpha, beta, T, omega, tol1, tol2, maxiter):
    X = T
    W = X
    Y = X
    iter0 = 1
    stop1 = 1
    stop2 = 1
    while stop1 > tol1 or stop2 > tol2:
        # the processing of computing W
        tran = (1/beta) * (Y+alpha*(T*omega))+X
        W = tran - (alpha/(alpha+beta))*omega*tran
        W[W < 0] = 0
        W[W > 1] = 1

        # the processing of computing X
        X_1 = svt(W-(1/beta)*Y, 1/beta)

        # the processing of computing Y
        Y = Y + beta*(X_1-W)

        stop1_0 = stop1
        if np.linalg.norm(X) != 0:
            stop1 = np.linalg.norm(X_1-X) / np.linalg.norm(X)
        else:
            stop1 = np.linalg.norm(X_1-X)
        stop2 = np.abs(stop1-stop1_0)/(max(1, np.abs(stop1_0)))
        X = X_1

        if iter0 >= maxiter:
            iter0 = maxiter
            print('reach maximum iteration,did not converge!')
            break
        iter0 = iter0 + 1
    T_recover = W
    return T_recover, iter0
def run_MC(t):
    # BNNR parameter
    maxiter = 300
    alpha = 1
    beta = 10
    tol1 = 2 * 1e-3
    tol2 = 1 * 1e-5
    omega = np.zeros(t.shape)
    omega[t.nonzero()] = 1
    WW, k = BNNR(alpha, beta, t, omega, tol1, tol2, maxiter)
    Smmi = WW[0:sm.shape[0], sm.shape[0]:WW.shape[1]]
    print(Smmi, k)
def list_flatten(l, a=None):
    a = list(a) if isinstance(a, (list, tuple)) else []
    for i in l:
        if isinstance(i, (list, tuple)):
            a = list_flatten(i, a)
        else:
            a.append(i)
    return a  
a=np.mat(np.zeros((541,831)))#Define a 541*831 matrix of zeros
AA=a.astype(int)#Convert matrix AA to integer
    smmi=np.array(AA.T)
    dt=pd.DataFrame(smmi)
    sm = read('myDrugSim16.xlsx')
    mi = read('myMiSim.xlsx')
    smmi_r = np.ones((sm.shape[0], mi.shape[0])) - smmi
    smmi_t = np.transpose(smmi)
    t0 = np.hstack((sm, smmi))
    t1 = np.hstack((smmi_t, mi))
    Tar = np.vstack((t0, t1))
    maxiter = 300
    alpha = 1
    beta = 10
    tol1 = 2 * 1e-3
    tol2 = 1 * 1e-5
    omega = np.zeros(Tar.shape)
    omega[Tar.nonzero()] = 1
    WW, k = BNNR(alpha, beta, Tar, omega, tol1, tol2, maxiter)
    Smmi = WW[0:sm.shape[0], sm.shape[0]:WW.shape[1]]
    print('The'+str(y+1)+'times ')
    #Calculate the score of elements that change from 1 to 0
    S1=Smmi.T[i,j]

          
































